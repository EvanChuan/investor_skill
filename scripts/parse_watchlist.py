#!/usr/bin/env python3
"""
parse_watchlist.py — 解析 TheMarketMemo Watchlist Excel 並輸出 Markdown 日報輔助檔

用法：
    python3 scripts/parse_watchlist.py [path/to/Market_Watchlist.xlsx]

若未指定路徑，自動尋找專案根目錄中最新的 Market Watchlist*.xlsx
輸出至 reports/watchlist_YYYY-MM-DD.md
"""

import sys
import os
import re
import glob
from datetime import datetime, date
import openpyxl


PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REPORTS_DIR = os.path.join(PROJECT_ROOT, "reports")


def find_latest_xlsx():
    pattern = os.path.join(PROJECT_ROOT, "Market Watchlist*.xlsx")
    files = glob.glob(pattern)
    if not files:
        raise FileNotFoundError("找不到 Market Watchlist*.xlsx，請確認檔案在專案根目錄")
    return max(files, key=os.path.getmtime)


def fmt_pct(val):
    if val is None:
        return "N/A"
    pct = val * 100
    arrow = "↑" if pct >= 0 else "↓"
    return f"{arrow} {abs(pct):.2f}%"


def fmt_rank(val):
    if val is None:
        return "—"
    return f"{val:.1f}"


def parse_sheet(ws, has_5d_20d=True):
    """
    解析工作表，回傳 (sections, all_rows) 的結構。
    sections: [ { name: str, rows: [ {ticker, name, price, d1, d5, d20, d60, r20, r60, r120, rank} ] } ]
    """
    # 欄位 index（0-based from column C = index 0 after stripping first 2 None cols）
    # C=0, D=1, E=2, F=3, G=4, H=5, I=6, J=7(sparkline/60), K=8, L=9, M=10, N=11
    # For 0518 (no 5D/20D/60D): C=0, D=1, E=2, F=3, J=4(sparkline), K=5, L=6, M=7, N=8

    sections = []
    current_section = None
    header_found = False

    for row in ws.iter_rows(min_row=1, values_only=True):
        vals = list(row)
        # Data starts at column C (index 2)
        c_val = vals[2] if len(vals) > 2 else None
        d_val = vals[3] if len(vals) > 3 else None
        e_val = vals[4] if len(vals) > 4 else None

        if c_val == "Ticker":
            header_found = True
            continue

        if not header_found:
            continue

        if c_val is None:
            continue

        # Section header: has ticker but no price
        if isinstance(e_val, str) or e_val is None:
            if d_val is None and c_val:
                # Section or sub-section header
                current_section = {"name": str(c_val), "rows": []}
                sections.append(current_section)
            continue

        if not isinstance(e_val, (int, float)):
            continue

        price = e_val
        d1 = vals[5] if len(vals) > 5 else None

        if has_5d_20d:
            d5  = vals[6]  if len(vals) > 6  else None
            d20 = vals[7]  if len(vals) > 7  else None
            d60 = vals[8]  if len(vals) > 8  else None
            # col J is sparkline, skip to K=index 10
            r20 = vals[10] if len(vals) > 10 else None
            r60 = vals[11] if len(vals) > 11 else None
            r120= vals[12] if len(vals) > 12 else None
            rank= vals[13] if len(vals) > 13 else None
        else:
            d5 = d20 = d60 = None
            r20 = vals[7]  if len(vals) > 7  else None
            r60 = vals[8]  if len(vals) > 8  else None
            r120= vals[9]  if len(vals) > 9  else None
            rank= vals[10] if len(vals) > 10 else None

        ticker = str(c_val).replace("BATS:", "")
        entry = {
            "ticker": ticker,
            "name":   str(d_val) if d_val else "",
            "price":  price,
            "d1":     d1,
            "d5":     d5,
            "d20":    d20,
            "d60":    d60,
            "r20":    r20,
            "r60":    r60,
            "r120":   r120,
            "rank":   rank,
        }

        if current_section is None:
            current_section = {"name": "其他", "rows": []}
            sections.append(current_section)

        current_section["rows"].append(entry)

    return sections


def rank_label(rank):
    if rank is None:
        return ""
    if rank >= 80:
        return "🔥"
    if rank >= 60:
        return "↑"
    if rank >= 40:
        return "→"
    return "↓"


def render_section_table(section, show_d5d20=True):
    rows = section["rows"]
    if not rows:
        return ""

    lines = [f"### {section['name']}"]
    if show_d5d20:
        lines.append("| 代號 | 價格 | 1D% | 5D% | 20D% | 60D% | Rank |")
        lines.append("|------|------|-----|-----|------|------|------|")
        for r in rows:
            label = rank_label(r["rank"])
            rank_str = f"{r['rank']:.1f} {label}" if r["rank"] is not None else "—"
            lines.append(
                f"| {r['ticker']} | ${r['price']:.2f} | {fmt_pct(r['d1'])} | "
                f"{fmt_pct(r['d5'])} | {fmt_pct(r['d20'])} | {fmt_pct(r['d60'])} | {rank_str} |"
            )
    else:
        lines.append("| 代號 | 價格 | 1D% | R20 | R60 | R120 | Rank |")
        lines.append("|------|------|-----|-----|-----|------|------|")
        for r in rows:
            label = rank_label(r["rank"])
            rank_str = f"{r['rank']:.1f} {label}" if r["rank"] is not None else "—"
            lines.append(
                f"| {r['ticker']} | ${r['price']:.2f} | {fmt_pct(r['d1'])} | "
                f"{fmt_rank(r['r20'])} | {fmt_rank(r['r60'])} | {fmt_rank(r['r120'])} | {rank_str} |"
            )
    return "\n".join(lines)


def top_n_by_rank(sections, n=10, min_rank=70):
    all_rows = []
    for s in sections:
        for r in s["rows"]:
            if r["rank"] is not None:
                r["_section"] = s["name"]
                all_rows.append(r)
    top = sorted(all_rows, key=lambda x: x["rank"], reverse=True)
    return [r for r in top[:n] if r["rank"] >= min_rank]


def detect_version(ws):
    """判斷是否含有 5D/20D/60D 欄位（0519 版本）"""
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        vals = list(row)
        if "5D%" in vals:
            return True
    return False


def build_report(xlsx_path, report_date):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    has_5d = detect_version(wb["Assets"])

    assets_sections    = parse_sheet(wb["Assets"],    has_5d_20d=has_5d)
    structure_sections = parse_sheet(wb["Structure"], has_5d_20d=has_5d)
    industry_sections  = parse_sheet(wb["Industry"],  has_5d_20d=has_5d)

    date_str = report_date.strftime("%Y-%m-%d")
    md_date  = report_date.strftime("%m/%d")

    lines = []
    lines.append(f"# Market Watchlist 分析 — {date_str}")
    lines.append(f"> 來源：TheMarketMemo Watchlist（{date_str} 收盤數據）")
    lines.append(f"> Rank = 0.2×R20 + 0.4×R60 + 0.4×R120（相對全市場百分位）")
    lines.append("")

    # ── 強勢族群 Top 10 ──
    lines.append("## 強勢板塊 Top 10（Industry Rank ≥ 70）")
    top_industry = top_n_by_rank(industry_sections, n=10, min_rank=70)
    if top_industry:
        lines.append("| 排名 | 代號 | 板塊 | 價格 | 1D% | Rank |")
        lines.append("|------|------|------|------|-----|------|")
        for i, r in enumerate(top_industry, 1):
            lines.append(
                f"| {i} | {r['ticker']} | {r['_section']} | ${r['price']:.2f} | "
                f"{fmt_pct(r['d1'])} | {r['rank']:.1f} 🔥 |"
            )
    else:
        lines.append("（本日無 Rank ≥ 70 的板塊）")
    lines.append("")

    # ── Assets 概覽 ──
    lines.append("## Assets — 全球市場概覽")
    lines.append("")
    for section in assets_sections:
        rendered = render_section_table(section, show_d5d20=has_5d)
        if rendered:
            lines.append(rendered)
            lines.append("")

    # ── Market Structure ──
    lines.append("## Market Structure — 市場結構")
    lines.append("")
    for section in structure_sections:
        rendered = render_section_table(section, show_d5d20=has_5d)
        if rendered:
            lines.append(rendered)
            lines.append("")

    # ── Industry/Thematic ──
    lines.append("## Industry/Thematic — 板塊與主題 ETF")
    lines.append("")
    for section in industry_sections:
        rendered = render_section_table(section, show_d5d20=has_5d)
        if rendered:
            lines.append(rendered)
            lines.append("")

    return "\n".join(lines)


def main():
    if len(sys.argv) > 1:
        xlsx_path = sys.argv[1]
        if not os.path.isabs(xlsx_path):
            xlsx_path = os.path.join(os.getcwd(), xlsx_path)
    else:
        xlsx_path = find_latest_xlsx()

    print(f"讀取：{xlsx_path}")

    # 從檔名提取日期 (MMDD) 或使用今天
    basename = os.path.basename(xlsx_path)
    match = re.search(r"(\d{4})", basename)
    if match:
        mmdd = match.group(1)
        year = date.today().year
        try:
            report_date = datetime.strptime(f"{year}{mmdd}", "%Y%m%d").date()
        except ValueError:
            report_date = date.today()
    else:
        report_date = date.today()

    os.makedirs(REPORTS_DIR, exist_ok=True)
    out_path = os.path.join(REPORTS_DIR, f"{report_date}_watchlist.md")

    report = build_report(xlsx_path, report_date)

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(report)

    print(f"輸出：{out_path}")


if __name__ == "__main__":
    main()
